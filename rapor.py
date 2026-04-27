"""
rapor.py — Generate Rapor Kurikulum Merdeka
Format: SD Taman Harapan 1 Bekasi
Satu halaman: info siswa → nilai + capaian kompetensi → ekskul → absensi → catatan → TTD
"""
from io import BytesIO
from datetime import date
import pandas as pd

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── Warna ──────────────────────────────────────────────────────────
WHITE  = colors.white
BLACK  = colors.black
GRAY   = colors.HexColor("#f5f5f5")
LGRAY  = colors.HexColor("#e0e0e0")
DBLUE  = colors.HexColor("#1a3a5c")

# ── Konstanta Sekolah ──────────────────────────────────────────────
NAMA_SEKOLAH   = "SD TAMAN HARAPAN"
NPSN           = "2023650"
ALAMAT_SEKOLAH = "Perum. Taman Harapan Baru, Kota Bekasi"
NAMA_KEPSEK    = "RAHMAT HIDAYAT, S.Psi., M.M."
NUPTK_KEPSEK   = "1642757659200010"

# ── Capaian Kompetensi Otomatis ────────────────────────────────────
CAPAIAN_TEMPLATE = {
    "A": "Sangat baik dalam memahami, menguasai, dan menerapkan konsep {mapel}. Mampu menganalisis dan menyajikan hasil belajar dengan sangat baik.",
    "B": "Baik dalam memahami dan menguasai konsep {mapel}. Mampu menerapkan pengetahuan dalam situasi yang relevan dengan baik.",
    "C": "Cukup baik dalam memahami konsep dasar {mapel}. Perlu peningkatan dalam penerapan dan pengembangan konsep lebih lanjut.",
    "D": "Mulai berkembang dalam memahami konsep {mapel}. Memerlukan bimbingan lebih lanjut untuk meningkatkan pemahaman dan keterampilan.",
}

def get_predikat(nilai):
    if nilai is None or (isinstance(nilai, float) and pd.isna(nilai)):
        return "-"
    if nilai >= 90: return "A"
    if nilai >= 80: return "B"
    if nilai >= 70: return "C"
    return "D"

def get_capaian_otomatis(mapel, nilai_p, nilai_k):
    avg = None
    vals = [v for v in [nilai_p, nilai_k] if v is not None and not (isinstance(v, float) and pd.isna(v))]
    if vals:
        avg = sum(vals) / len(vals)
    pred = get_predikat(avg)
    template = CAPAIAN_TEMPLATE.get(pred, CAPAIAN_TEMPLATE["C"])
    return template.format(mapel=mapel)

def fmt_nilai(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "-"
    return str(int(round(v)))


def generate_rapor_pdf(
    nama_siswa: str,
    kelas: str,
    semester: str,
    tahun_ajar: str,
    nilai_df: pd.DataFrame,       # cols: mapel, pengetahuan, keterampilan, sikap, capaian
    nama_wali_kelas: str = "",
    nuptk_wali: str = "",
    fase: str = "",
    nis: str = "",
    nisn: str = "",
    alamat_siswa: str = "",
    ekskul_df: pd.DataFrame = None,
    absen_count: dict = None,     # {"S":0,"I":0,"A":0}
    catatan_wali: str = "",
    tanggapan_ortu: str = "",
    nama_kepala: str = NAMA_KEPSEK,
    nuptk_kepala: str = NUPTK_KEPSEK,
) -> bytes:

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.2*cm, bottomMargin=1.2*cm
    )

    def sty(name, size=9, bold=False, align=TA_LEFT, color=BLACK, leading=None):
        return ParagraphStyle(
            name,
            fontName="Helvetica-Bold" if bold else "Helvetica",
            fontSize=size,
            textColor=color,
            alignment=align,
            leading=leading or (size * 1.4),
            spaceAfter=0,
            spaceBefore=0,
        )

    W = A4[0] - 3*cm  # usable width

    story = []

    # ══ HEADER ════════════════════════════════════════════════════
    header_data = [[
        Paragraph("LAPORAN HASIL BELAJAR MURID", sty("h1", 12, True, TA_CENTER)),
    ],[
        Paragraph(f"(RAPOR)", sty("h2", 10, False, TA_CENTER)),
    ]]
    hdr_tbl = Table(header_data, colWidths=[W])
    hdr_tbl.setStyle(TableStyle([
        ("ALIGN",       (0,0),(-1,-1), "CENTER"),
        ("TOPPADDING",  (0,0),(-1,-1), 2),
        ("BOTTOMPADDING",(0,0),(-1,-1), 2),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 6))

    # ══ INFO SISWA ════════════════════════════════════════════════
    # Mapping semester → nomor
    sem_num = "1 (Satu)" if "Ganjil" in semester else "2 (Dua)"
    # Fase otomatis dari kelas jika tidak diisi
    if not fase:
        kelas_num = kelas.strip()[:2].strip()
        if kelas_num in ["I","II"]:       fase = "A"
        elif kelas_num in ["III","IV"]:   fase = "B"
        else:                              fase = "C"

    info_left = [
        ["Nama Murid", ":", nama_siswa],
        ["NIS/NISN",   ":", f"{nis} / {nisn}" if (nis or nisn) else "-"],
        ["Sekolah",    ":", NAMA_SEKOLAH],
        ["Alamat",     ":", alamat_siswa or ALAMAT_SEKOLAH],
    ]
    info_right = [
        ["Kelas",         ":", kelas],
        ["Fase",          ":", fase],
        ["Semester",      ":", sem_num],
        ["Tahun Ajaran",  ":", tahun_ajar],
    ]

    def info_table(data):
        tbl = Table(data, colWidths=[2.8*cm, 0.4*cm, None])
        tbl.setStyle(TableStyle([
            ("FONTNAME",     (0,0),(-1,-1), "Helvetica"),
            ("FONTSIZE",     (0,0),(-1,-1), 8.5),
            ("FONTNAME",     (0,0),(0,-1),  "Helvetica"),
            ("TOPPADDING",   (0,0),(-1,-1), 2),
            ("BOTTOMPADDING",(0,0),(-1,-1), 2),
            ("VALIGN",       (0,0),(-1,-1), "TOP"),
            ("BOX",          (0,0),(-1,-1), 0.5, LGRAY),
            ("INNERGRID",    (0,0),(-1,-1), 0.3, LGRAY),
        ]))
        return tbl

    half = W/2 - 0.2*cm
    info_tbl = Table(
        [[info_table(info_left), info_table(info_right)]],
        colWidths=[half+0.4*cm, half]
    )
    info_tbl.setStyle(TableStyle([
        ("VALIGN",  (0,0),(-1,-1), "TOP"),
        ("LEFTPADDING",  (1,0),(1,0), 4),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 8))

    # ══ TABEL NILAI + CAPAIAN KOMPETENSI ═════════════════════════
    # Header
    nilai_header = [
        [
            Paragraph("No.", sty("th", 8, True, TA_CENTER)),
            Paragraph("Mata Pelajaran", sty("th", 8, True, TA_CENTER)),
            Paragraph("Nilai\nAkhir", sty("th", 8, True, TA_CENTER)),
            Paragraph("Capaian Kompetensi", sty("th", 8, True, TA_CENTER)),
        ]
    ]

    nilai_rows = []
    if nilai_df is None or nilai_df.empty:
        nilai_rows.append([
            Paragraph("", sty("td",8)),
            Paragraph("Belum ada data nilai", sty("td",8)),
            Paragraph("", sty("td",8,False,TA_CENTER)),
            Paragraph("", sty("td",8)),
        ])
    else:
        for i, (_, r) in enumerate(nilai_df.iterrows(), 1):
            mapel    = str(r.get("mapel",""))
            p_val    = r.get("pengetahuan")
            k_val    = r.get("keterampilan")
            # Nilai akhir = rata-rata P dan K
            vals = [v for v in [p_val, k_val] if v is not None and not (isinstance(v, float) and pd.isna(v))]
            nilai_akhir = round(sum(vals)/len(vals)) if vals else None
            # Capaian
            capaian_custom = str(r.get("capaian","")).strip()
            capaian = capaian_custom if capaian_custom else get_capaian_otomatis(mapel, p_val, k_val)

            nilai_rows.append([
                Paragraph(str(i), sty("td",8,False,TA_CENTER)),
                Paragraph(mapel, sty("td",8)),
                Paragraph(fmt_nilai(nilai_akhir), sty("td",8,True,TA_CENTER)),
                Paragraph(capaian, sty("td",8)),
            ])

    all_rows = nilai_header + nilai_rows
    col_w    = [0.8*cm, 3.8*cm, 1.5*cm, W - 0.8*cm - 3.8*cm - 1.5*cm]
    nilai_tbl = Table(all_rows, colWidths=col_w, repeatRows=1)
    nilai_tbl.setStyle(TableStyle([
        # Header
        ("BACKGROUND",    (0,0),(-1,0), LGRAY),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0), 8),
        ("ALIGN",         (0,0),(-1,0), "CENTER"),
        ("VALIGN",        (0,0),(-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,0), 4),
        ("BOTTOMPADDING", (0,0),(-1,0), 4),
        # Body
        ("FONTNAME",      (0,1),(-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1),(-1,-1), 8),
        ("VALIGN",        (0,1),(-1,-1), "TOP"),
        ("TOPPADDING",    (0,1),(-1,-1), 4),
        ("BOTTOMPADDING", (0,1),(-1,-1), 14),
        # Grid
        ("BOX",           (0,0),(-1,-1), 0.5, BLACK),
        ("INNERGRID",     (0,0),(-1,-1), 0.3, LGRAY),
    ]))
    story.append(nilai_tbl)
    story.append(Spacer(1, 10))

    # ══ EKSKUL ═══════════════════════════════════════════════════
    ekskul_header = [[
        Paragraph("No.", sty("th",8,True,TA_CENTER)),
        Paragraph("Ekstrakurikuler", sty("th",8,True,TA_CENTER)),
        Paragraph("Keterangan", sty("th",8,True,TA_CENTER)),
    ]]
    ekskul_rows = []
    if ekskul_df is not None and not ekskul_df.empty:
        for i, (_, e) in enumerate(ekskul_df.iterrows(), 1):
            ekskul_rows.append([
                Paragraph(str(i), sty("td",8,False,TA_CENTER)),
                Paragraph(str(e.get("nama_ekskul","")), sty("td",8)),
                Paragraph(str(e.get("keterangan","")), sty("td",8)),
            ])
    else:
        for i, nama in enumerate(["Pramuka","Komputer","Anak Beriman dan Berkepribadian"], 1):
            ekskul_rows.append([
                Paragraph(str(i), sty("td",8,False,TA_CENTER)),
                Paragraph(nama, sty("td",8)),
                Paragraph("", sty("td",8)),
            ])

    ekskul_tbl = Table(ekskul_header + ekskul_rows,
                       colWidths=[0.8*cm, 5*cm, W-0.8*cm-5*cm], repeatRows=1)
    ekskul_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), LGRAY),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("ALIGN",         (0,0),(0,-1), "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 14),
        ("BOX",           (0,0),(-1,-1), 0.5, BLACK),
        ("INNERGRID",     (0,0),(-1,-1), 0.3, LGRAY),
    ]))
    story.append(ekskul_tbl)
    story.append(Spacer(1, 8))

    # ══ KETIDAKHADIRAN + CATATAN WALI KELAS ═══════════════════════
    ab = absen_count or {"S":0,"I":0,"A":0}
    absen_data = [
        [Paragraph("Ketidakhadiran", sty("th",8,True,TA_CENTER)), "", Paragraph("Catatan Wali Kelas", sty("th",8,True,TA_CENTER))],
        [Paragraph(f"Sakit              {ab.get('S',0)} hari", sty("td",8)), "",
         Paragraph(catatan_wali or "-", sty("td",8))],
        [Paragraph(f"Izin               {ab.get('I',0)} hari", sty("td",8)), "", ""],
        [Paragraph(f"Tanpa Keterangan   {ab.get('A',0)} hari", sty("td",8)), "", ""],
    ]
    half_ab = W/2 - 0.3*cm
    absen_tbl = Table(absen_data,
                      colWidths=[half_ab, 0.6*cm, half_ab],
                      rowHeights=[0.6*cm, None, None, None])
    absen_tbl.setStyle(TableStyle([
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("BACKGROUND",    (0,0),(0,0), LGRAY),
        ("BACKGROUND",    (2,0),(2,0), LGRAY),
        ("FONTNAME",      (0,0),(0,0), "Helvetica-Bold"),
        ("FONTNAME",      (2,0),(2,0), "Helvetica-Bold"),
        ("ALIGN",         (0,0),(0,0), "CENTER"),
        ("ALIGN",         (2,0),(2,0), "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING",   (0,0),(-1,-1), 4),
        ("BOX",           (0,0),(0,-1), 0.5, BLACK),
        ("BOX",           (2,0),(2,-1), 0.5, BLACK),
        ("INNERGRID",     (0,0),(0,-1), 0.3, LGRAY),
        ("SPAN",          (2,1),(2,3)),
    ]))
    story.append(absen_tbl)
    story.append(Spacer(1, 6))

    # ══ TANGGAPAN ORANG TUA ═══════════════════════════════════════
    ortu_data = [
        [Paragraph("Tanggapan Orang Tua / Wali Murid", sty("th",8,True,TA_CENTER))],
        [Paragraph(tanggapan_ortu or " ", sty("td",8))],
    ]
    ortu_tbl = Table(ortu_data, colWidths=[W], rowHeights=[0.6*cm, 1.8*cm])
    ortu_tbl.setStyle(TableStyle([
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("BACKGROUND",    (0,0),(0,0), LGRAY),
        ("FONTNAME",      (0,0),(0,0), "Helvetica-Bold"),
        ("ALIGN",         (0,0),(0,0), "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING",   (0,1),(0,1), 4),
        ("BOX",           (0,0),(-1,-1), 0.5, BLACK),
        ("INNERGRID",     (0,0),(-1,-1), 0.3, LGRAY),
    ]))
    story.append(ortu_tbl)
    story.append(Spacer(1, 12))

    # ══ TANDA TANGAN ══════════════════════════════════════════════
    tgl_str = f"Kota Bekasi, {date.today().strftime('%d %B %Y')}"
    # Nomor TTD: kiri=ortu, tengah=kepsek, kanan=wali kelas
    ttd_data = [[
        Paragraph("Orang Tua Murid", sty("td",8,False,TA_LEFT)),
        Paragraph(tgl_str, sty("td",8,False,TA_RIGHT)),
    ],[
        Paragraph("", sty("td",8)),
        Paragraph("Wali Kelas", sty("td",8,False,TA_RIGHT)),
    ]]
    ttd_top = Table(ttd_data, colWidths=[W/2, W/2])
    ttd_top.setStyle(TableStyle([
        ("FONTSIZE", (0,0),(-1,-1), 8),
        ("VALIGN",   (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",(0,0),(-1,-1), 0),
    ]))
    story.append(ttd_top)

    # Ruang TTD
    story.append(Spacer(1, 1.8*cm))

    # Nama TTD
    ttd_nama = Table([[
        Paragraph("___________________________", sty("td",8,False,TA_LEFT)),
        Paragraph(f"<b>{nama_wali_kelas.upper() if nama_wali_kelas else '___________________________'}</b>",
                  sty("td",8,True,TA_RIGHT)),
    ],[
        Paragraph("", sty("td",8)),
        Paragraph(f"NUPTK {nuptk_wali}" if nuptk_wali else "", sty("td",8,False,TA_RIGHT)),
    ]], colWidths=[W/2, W/2])
    ttd_nama.setStyle(TableStyle([
        ("FONTSIZE", (0,0),(-1,-1), 8),
        ("TOPPADDING",(0,0),(-1,-1), 1),
    ]))
    story.append(ttd_nama)

    story.append(Spacer(1, 14))

    # Kepala Sekolah — tengah
    kepsek_data = [[
        Paragraph("", sty("td",8)),
        Paragraph("Kepala Sekolah", sty("td",8,False,TA_CENTER)),
        Paragraph("", sty("td",8)),
    ],[
        Paragraph("", sty("td",8)),
        Paragraph("", sty("td",8)),
        Paragraph("", sty("td",8)),
    ],[
        Paragraph("", sty("td",8)),
        Paragraph("", sty("td",8)),
        Paragraph("", sty("td",8)),
    ],[
        Paragraph("", sty("td",8)),
        Paragraph(f"<b>{nama_kepala.upper()}</b>", sty("td",8,True,TA_CENTER)),
        Paragraph("", sty("td",8)),
    ],[
        Paragraph("", sty("td",8)),
        Paragraph(f"NUPTK {nuptk_kepala}", sty("td",8,False,TA_CENTER)),
        Paragraph("", sty("td",8)),
    ]]
    kepsek_tbl = Table(kepsek_data, colWidths=[W*0.25, W*0.5, W*0.25],
                       rowHeights=[0.4*cm, 0.4*cm, 0.4*cm, 0.4*cm, 0.4*cm])
    kepsek_tbl.setStyle(TableStyle([
        ("FONTSIZE",    (0,0),(-1,-1), 8),
        ("VALIGN",      (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",  (0,0),(-1,-1), 1),
    ]))
    story.append(kepsek_tbl)

    doc.build(story)
    return buf.getvalue()


def generate_rekap_excel(
    kelas: str,
    semester: str,
    tahun_ajar: str,
    nilai_df: pd.DataFrame,
    siswa_list: list,
) -> bytes:
    """
    Generate Excel rekap nilai persis seperti foto:
    Header: DAFTAR NILAI RAPOR SEMESTER X, KELAS, SD TAMAN HARAPAN
    Kolom: NO, NAMA SISWA, per-mapel (nilai akhir), JMLH, RT, RANK
    """
    buf = BytesIO()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Rekap {semester}"

    # ── Style helpers ──────────────────────────────────────────────
    BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
    LBLUE_FILL  = PatternFill("solid", fgColor="2E75B6")
    LLBLUE_FILL = PatternFill("solid", fgColor="BDD7EE")
    GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
    ALT_FILL    = PatternFill("solid", fgColor="DEEAF1")

    thin  = Side(style="thin",   color="000000")
    thick = Side(style="medium", color="000000")
    bord_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
    bord_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    def hdr(cell, text, size=11, bold=True, color="FFFFFF", bg=None, align="center", wrap=False):
        cell.value = text
        cell.font  = Font(name="Calibri", size=size, bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   wrap_text=wrap)
        cell.border = bord_all
        if bg: cell.fill = PatternFill("solid", fgColor=bg)

    def dat(cell, value, bold=False, color="000000", align="center", bg=None):
        cell.value = value
        cell.font  = Font(name="Calibri", size=10, bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = bord_all
        if bg: cell.fill = PatternFill("solid", fgColor=bg)

    # ── Mapel list ─────────────────────────────────────────────────
    mapel_list = sorted(nilai_df["mapel"].dropna().unique().tolist()) if not nilai_df.empty else []
    n_mapel    = len(mapel_list)

    # Column layout:
    # A=NO, B=NAMA SISWA, C...(C+n_mapel-1)=mapel, then JMLH, RT, RANK
    col_no   = 1
    col_nama = 2
    col_mp   = 3
    col_jmlh = col_mp + n_mapel
    col_rt   = col_jmlh + 1
    col_rank = col_rt + 1
    total_col = col_rank

    # ── Row 1: Judul ───────────────────────────────────────────────
    sem_label = "1 (SATU)" if "Ganjil" in semester else "2 (DUA)"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)
    c = ws.cell(1, 1, f"DAFTAR NILAI RAPOR SEMESTER {sem_label}")
    c.font = Font(name="Calibri", size=14, bold=True, color="000000")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_col)
    c2 = ws.cell(2, 1, f"KELAS {kelas.upper()}")
    c2.font = Font(name="Calibri", size=12, bold=True, color="000000")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_col)
    c3 = ws.cell(3, 1, f"{NAMA_SEKOLAH} TAHUN AJARAN {tahun_ajar}")
    c3.font = Font(name="Calibri", size=12, bold=True, color="000000")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # ── Row 4: blank spacer ────────────────────────────────────────
    ws.row_dimensions[4].height = 6

    # ── Row 5: Header row 1 ────────────────────────────────────────
    # NO — merge rows 5-6
    ws.merge_cells(start_row=5, start_column=col_no, end_row=6, end_column=col_no)
    hdr(ws.cell(5, col_no, "NO"), "NO", size=10, bg="1F4E79")

    # NAMA SISWA — merge rows 5-6
    ws.merge_cells(start_row=5, start_column=col_nama, end_row=6, end_column=col_nama)
    hdr(ws.cell(5, col_nama, "NAMA SISWA"), "NAMA SISWA", size=10, bg="1F4E79")

    # MATA PELAJARAN — merge across all mapel cols, rows 5 only
    if n_mapel > 0:
        ws.merge_cells(start_row=5, start_column=col_mp, end_row=5, end_column=col_mp+n_mapel-1)
        hdr(ws.cell(5, col_mp, "MATA PELAJARAN"), "MATA PELAJARAN", size=10, bg="2E75B6")

    # JMLH, RT, RANK — merge rows 5-6
    for ci, lbl in [(col_jmlh,"JMLH"),(col_rt,"RT"),(col_rank,"RANK")]:
        ws.merge_cells(start_row=5, start_column=ci, end_row=6, end_column=ci)
        hdr(ws.cell(5, ci, lbl), lbl, size=10, bg="1F4E79")

    ws.row_dimensions[5].height = 18

    # ── Row 6: Mapel abbreviations ─────────────────────────────────
    ABBR = {
        "Pend. Agama & Budi Pekerti":"PABP","Pendidikan Agama":"PABP",
        "PADB":"PABP","PAI":"PABP",
        "PPKn":"PANC","Pancasila":"PANC",
        "Bahasa Indonesia":"BIND","B. Indonesia":"BIND",
        "Matematika":"MTK",
        "IPAS":"IPAS","IPA":"IPA","IPS":"IPS",
        "PJOK":"PJOK",
        "Seni Budaya":"SBK","SBdP":"SBK","Seni":"SBK",
        "Bahasa Inggris":"BING","B. Inggris":"BING",
        "B. Arab":"KKA","Bahasa Arab":"KKA",
        "TIK":"BJW",
        "Tahfidh":"MLK2",
        "Mulok":"MLK3","Muatan Lokal":"MLK3",
    }
    for i, mp in enumerate(mapel_list):
        abbr = ABBR.get(mp, mp[:4].upper())
        hdr(ws.cell(6, col_mp+i, abbr), abbr, size=9, bg="2E75B6")
    ws.row_dimensions[6].height = 16

    # ── Row 7: Filter row (like Excel table style) ─────────────────
    ws.row_dimensions[7].height = 14
    for ci in range(1, total_col+1):
        c = ws.cell(7, ci)
        c.fill = LBLUE_FILL
        c.border = bord_all

    # Number rows for filter markers
    for ci, val in [(col_no,1),(col_nama,2)]:
        dat(ws.cell(7, ci), val, bg="BDD7EE", color="000000")
    for i in range(n_mapel):
        dat(ws.cell(7, col_mp+i), col_mp+i, bg="BDD7EE", color="000000")
    for ci, val in [(col_jmlh,col_jmlh),(col_rt,col_rt),(col_rank,col_rank)]:
        dat(ws.cell(7, ci), val, bg="BDD7EE", color="000000")

    # ── Data rows ──────────────────────────────────────────────────
    # Build nilai lookup: siswa_id → {mapel: nilai_akhir}
    nilai_lookup = {}
    if not nilai_df.empty:
        for _, r in nilai_df.iterrows():
            sid  = r.get("siswa_id")
            mp   = r.get("mapel","")
            p    = r.get("pengetahuan")
            k    = r.get("keterampilan")
            vals = [v for v in [p,k] if v is not None and not (isinstance(v, float) and pd.isna(v))]
            na   = round(sum(vals)/len(vals)) if vals else 0
            if sid not in nilai_lookup: nilai_lookup[sid] = {}
            nilai_lookup[sid][mp] = na

    # Hitung total per siswa untuk ranking
    siswa_totals = []
    for sid, sname in siswa_list:
        mp_vals = [nilai_lookup.get(sid,{}).get(mp, 0) for mp in mapel_list]
        total   = sum(mp_vals)
        rt      = round(total/n_mapel, 2) if n_mapel > 0 else 0
        siswa_totals.append((sid, sname, mp_vals, total, rt))

    # Ranking berdasarkan total (descending)
    sorted_by_rank = sorted(siswa_totals, key=lambda x: x[3], reverse=True)
    rank_map = {}
    for rank_i, (sid,_,_,_,_) in enumerate(sorted_by_rank, 1):
        rank_map[sid] = rank_i

    data_start = 8
    for idx, (sid, sname, mp_vals, total, rt) in enumerate(siswa_totals):
        row = data_start + idx
        bg  = "FFFFFF" if idx % 2 == 0 else "DEEAF1"

        dat(ws.cell(row, col_no),   idx+1, bg=bg)
        dat(ws.cell(row, col_nama), sname, align="left", bg=bg)

        for i, val in enumerate(mp_vals):
            dat(ws.cell(row, col_mp+i), val if val else 0, bg=bg)

        dat(ws.cell(row, col_jmlh), total, bold=True, bg=bg)
        dat(ws.cell(row, col_rt),   rt, bold=True, bg=bg)
        dat(ws.cell(row, col_rank), rank_map[sid], bold=True, bg=bg)

        ws.row_dimensions[row].height = 15

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(col_no)].width   = 4
    ws.column_dimensions[get_column_letter(col_nama)].width = 28
    for i in range(n_mapel):
        ws.column_dimensions[get_column_letter(col_mp+i)].width = 6
    ws.column_dimensions[get_column_letter(col_jmlh)].width = 7
    ws.column_dimensions[get_column_letter(col_rt)].width   = 7
    ws.column_dimensions[get_column_letter(col_rank)].width = 6

    # ── Freeze panes ───────────────────────────────────────────────
    ws.freeze_panes = ws.cell(data_start, col_mp)

    wb.save(buf)
    return buf.getvalue()



def generate_rekap_tp_excel(
    kelas: str,
    semester: str,
    tahun_ajar: str,
    mapel: str,
    nilai_df: pd.DataFrame,   # cols: siswa_id, siswa, tp1..tp10, asts, asas, nr, capaian_maksimal
    siswa_list: list,
) -> bytes:
    """
    Generate Excel rekap nilai TP per mapel — format persis seperti foto:
    NO | NAMA | TP1..TP10 | ASTS | ASAS | NR | Capaian Maksimal
    """
    buf = BytesIO()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = mapel[:28]

    BLUE  = PatternFill("solid", fgColor="1F4E79")
    LBLUE = PatternFill("solid", fgColor="2E75B6")
    LLBLU = PatternFill("solid", fgColor="BDD7EE")
    thin  = Side(style="thin", color="000000")
    bord  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, text, bg=None, color="FFFFFF", size=10, bold=True, align="center"):
        cell.value = text
        cell.font  = Font(name="Calibri", size=size, bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = bord
        if bg: cell.fill = PatternFill("solid", fgColor=bg)

    def dat(cell, val, bold=False, color="000000", align="center", bg=None):
        cell.value = val
        cell.font  = Font(name="Calibri", size=10, bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = bord
        if bg: cell.fill = PatternFill("solid", fgColor=bg)

    sem_label = "1 (SATU)" if "Ganjil" in semester else "2 (DUA)"
    last_col  = 17  # NO, NAMA, TP1-10, ASTS, ASAS, NR, Capaian = 17 cols

    # ── Title rows ────────────────────────────────────────────────
    for r, txt, sz in [
        (1, f"DAFTAR NILAI — {mapel.upper()}", 13),
        (2, f"SEMESTER {sem_label}  |  KELAS {kelas.upper()}", 11),
        (3, f"{NAMA_SEKOLAH}  |  TAHUN AJARAN {tahun_ajar}", 11),
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        c = ws.cell(r, 1, txt)
        c.font = Font(name="Calibri", size=sz, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[4].height = 6  # spacer

    # ── Header row 5 — merge NO and NAMA across rows 5-6 ─────────
    for ci, lbl in [(1,"NO"),(2,"NAMA SISWA")]:
        ws.merge_cells(start_row=5, start_column=ci, end_row=6, end_column=ci)
        hdr(ws.cell(5, ci, lbl), lbl, bg="1F4E79")

    # TP header merge
    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=12)
    hdr(ws.cell(5, 3, "TUJUAN PEMBELAJARAN"), "TUJUAN PEMBELAJARAN", bg="2E75B6")

    # ASTS, ASAS, NR — merge rows 5-6
    for ci, lbl in [(13,"ASTS"),(14,"ASAS"),(15,"NR")]:
        ws.merge_cells(start_row=5, start_column=ci, end_row=6, end_column=ci)
        hdr(ws.cell(5, ci, lbl), lbl, bg="1F4E79")

    # Capaian Maksimal — merge rows 5-6
    ws.merge_cells(start_row=5, start_column=16, end_row=6, end_column=last_col)
    hdr(ws.cell(5, 16, "Capaian Maksimal"), "Capaian Maksimal", bg="1F4E79")

    # TP1-TP10 header row 6
    for i in range(10):
        hdr(ws.cell(6, 3+i, f"TP{i+1}"), f"TP{i+1}", bg="2E75B6", size=9)

    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 14

    # Filter row 7
    for ci in range(1, last_col+1):
        c = ws.cell(7, ci)
        c.fill = LLBLU
        c.border = bord
        dat(ws.cell(7, ci), ci, bg="BDD7EE", color="1F4E79")

    ws.row_dimensions[7].height = 13

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    for i in range(10): ws.column_dimensions[get_column_letter(3+i)].width = 6
    ws.column_dimensions[get_column_letter(13)].width = 7
    ws.column_dimensions[get_column_letter(14)].width = 7
    ws.column_dimensions[get_column_letter(15)].width = 7
    ws.column_dimensions[get_column_letter(16)].width = 40

    # Build lookup by siswa_id
    lookup = {}
    if not nilai_df.empty:
        for _, r in nilai_df.iterrows():
            lookup[int(r["siswa_id"])] = r

    # ── Data rows ─────────────────────────────────────────────────
    data_row = 8
    for idx, (sid, sname) in enumerate(siswa_list, 1):
        bg_hex = "FFFFFF" if idx % 2 == 1 else "DEEAF1"
        bg = PatternFill("solid", fgColor=bg_hex)
        row = lookup.get(sid, {})

        dat(ws.cell(data_row, 1), idx, bg=bg_hex)
        dat(ws.cell(data_row, 2), sname, align="left", bg=bg_hex)

        for i, col_key in enumerate(['tp1','tp2','tp3','tp4','tp5','tp6','tp7','tp8','tp9','tp10']):
            v = row.get(col_key)
            try: v = int(float(v)) if v is not None and str(v) not in ("","nan","None") else None
            except: v = None
            dat(ws.cell(data_row, 3+i), v if v is not None else "", bg=bg_hex)

        for ci, col_key in [(13,'asts'),(14,'asas')]:
            v = row.get(col_key)
            try: v = int(float(v)) if v is not None and str(v) not in ("","nan","None") else None
            except: v = None
            dat(ws.cell(data_row, ci), v if v is not None else "", bg=bg_hex)

        nr = row.get("nr")
        try: nr = round(float(nr), 2) if nr is not None and str(nr) not in ("","nan","None") else ""
        except: nr = ""
        dat(ws.cell(data_row, 15), nr, bold=True, bg=bg_hex,
            color="1F4E79" if nr and float(nr) >= 70 else "C0281E" if nr else "000000")

        capaian = row.get("capaian_maksimal", "") or ""
        dat(ws.cell(data_row, 16), capaian, align="left", bg=bg_hex)

        ws.row_dimensions[data_row].height = 15
        data_row += 1

    # Footer
    ws.cell(data_row+1, 1, f"Dicetak: {date.today().strftime('%d/%m/%Y')}").font = \
        Font(name="Calibri", size=8, color="9AA0B8")

    ws.freeze_panes = ws.cell(8, 3)
    wb.save(buf)
    return buf.getvalue()


# ── CETAK ABSENSI PDF ─────────────────────────────────────────────

def generate_absen_guru_pdf(
    bulan: int,
    tahun: int,
    df_harian: pd.DataFrame,   # cols: nama, kelas, tanggal, status
    df_rekap:  pd.DataFrame,   # cols: nama, kelas, hadir, alpa, izin, sakit, total_hari
) -> bytes:
    """
    Generate PDF absensi guru per bulan — A4 landscape.
    Halaman 1: Tabel harian (nama × tanggal).
    Halaman 2: Rekapitulasi (nama, total H/A/I/S).
    """
    import calendar as cal_mod
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import PageBreak

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.2*cm, rightMargin=1.2*cm,
        topMargin=1.2*cm, bottomMargin=1.2*cm
    )

    nama_bulan = cal_mod.month_name[bulan]
    jumlah_hari = cal_mod.monthrange(tahun, bulan)[1]

    def sty(name, size=8, bold=False, align=TA_CENTER, color=BLACK):
        return ParagraphStyle(name, fontName="Helvetica-Bold" if bold else "Helvetica",
                              fontSize=size, textColor=color, alignment=align,
                              leading=size*1.3, spaceAfter=0, spaceBefore=0)

    W_L = landscape(A4)[0] - 2.4*cm  # usable width landscape

    thin  = colors.HexColor("#cccccc")
    hdr_c = colors.HexColor("#1F4E79")
    alt_c = colors.HexColor("#EBF5FB")

    story = []

    # ── HALAMAN 1: TABEL HARIAN ───────────────────────────────────
    story.append(Paragraph(f"DAFTAR HADIR GURU — {nama_bulan.upper()} {tahun}", sty("T", 11, True)))
    story.append(Paragraph(f"{NAMA_SEKOLAH} | NPSN {NPSN}", sty("S", 9)))
    story.append(Spacer(1, 6))

    # Build pivot: nama × tanggal
    status_map = {}
    if not df_harian.empty:
        for _, r in df_harian.iterrows():
            key = (r.get("nama",""), str(r.get("tanggal","")))
            status_map[key] = r.get("status","")

    guru_list = df_rekap["nama"].tolist() if not df_rekap.empty else []

    # Column widths: nama(3cm) + kelas(3cm) + 31 hari(0.55cm each)
    col_w_h = [3*cm, 2.5*cm] + [0.55*cm]*jumlah_hari + [0.7*cm, 0.7*cm, 0.7*cm, 0.7*cm]
    
    hdr_row = [
        Paragraph("Nama Guru", sty("h",7,True,TA_LEFT,colors.white)),
        Paragraph("Kelas", sty("h",7,True,TA_CENTER,colors.white)),
    ] + [
        Paragraph(str(d), sty("h",6,True,TA_CENTER,colors.white))
        for d in range(1, jumlah_hari+1)
    ] + [
        Paragraph("H", sty("h",7,True,TA_CENTER,colors.white)),
        Paragraph("A", sty("h",7,True,TA_CENTER,colors.white)),
        Paragraph("I", sty("h",7,True,TA_CENTER,colors.white)),
        Paragraph("S", sty("h",7,True,TA_CENTER,colors.white)),
    ]

    rows_h = [hdr_row]
    for idx, (_, rec) in enumerate(df_rekap.iterrows()):
        nama  = rec.get("nama","")
        kelas = str(rec.get("kelas",""))[:12]
        row   = [
            Paragraph(nama.split(",")[0][:20], sty("d",6,False,TA_LEFT)),
            Paragraph(kelas, sty("d",6,False,TA_CENTER)),
        ]
        for d in range(1, jumlah_hari+1):
            tgl_str = f"{tahun}-{str(bulan).zfill(2)}-{str(d).zfill(2)}"
            st = status_map.get((nama, tgl_str), "")
            color = {"H":colors.HexColor("#1E6B2E"),"A":colors.HexColor("#C0281E"),
                     "I":colors.HexColor("#8A6000"),"S":colors.HexColor("#1A4A9A")}.get(st, BLACK)
            row.append(Paragraph(st, sty("d",6,False,TA_CENTER,color)))
        row += [
            Paragraph(str(int(rec.get("hadir",0))),  sty("d",6,True,TA_CENTER,colors.HexColor("#1E6B2E"))),
            Paragraph(str(int(rec.get("alpa",0))),   sty("d",6,True,TA_CENTER,colors.HexColor("#C0281E"))),
            Paragraph(str(int(rec.get("izin",0))),   sty("d",6,True,TA_CENTER,colors.HexColor("#8A6000"))),
            Paragraph(str(int(rec.get("sakit",0))),  sty("d",6,True,TA_CENTER,colors.HexColor("#1A4A9A"))),
        ]
        rows_h.append(row)

    tbl_h = Table(rows_h, colWidths=col_w_h, repeatRows=1)
    tbl_h.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  hdr_c),
        ("FONTSIZE",      (0,0),(-1,-1), 7),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 2),
        ("BOTTOMPADDING", (0,0),(-1,-1), 2),
        ("BOX",           (0,0),(-1,-1), 0.5, thin),
        ("INNERGRID",     (0,0),(-1,-1), 0.3, thin),
        *[("BACKGROUND",(0,i),(-1,i), alt_c) for i in range(2,len(rows_h),2)],
    ]))
    story.append(tbl_h)

    # ── HALAMAN 2: REKAPITULASI ───────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph(f"REKAPITULASI KEHADIRAN GURU — {nama_bulan.upper()} {tahun}", sty("T2",11,True)))
    story.append(Paragraph(f"{NAMA_SEKOLAH} | NPSN {NPSN}", sty("S2",9)))
    story.append(Spacer(1, 8))

    rkp_cols = [1*cm, 5*cm, 3*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2*cm, 2*cm]
    rkp_hdr  = [Paragraph(t, sty("rh",8,True,TA_CENTER,colors.white))
                for t in ["No","Nama Guru","Kelas","Hadir","Alpa","Izin","Sakit","Total Hari","% Hadir"]]
    rkp_rows = [rkp_hdr]
    for idx, (_, r) in enumerate(df_rekap.iterrows(), 1):
        total = r.get("total_hari",0) or 1
        h     = int(r.get("hadir",0))
        pct   = f"{h/total*100:.1f}%"
        rkp_rows.append([
            Paragraph(str(idx),              sty("rd",8,False,TA_CENTER)),
            Paragraph(r.get("nama","").split(",")[0], sty("rd",8,False,TA_LEFT)),
            Paragraph(str(r.get("kelas",""))[:15], sty("rd",8,False,TA_LEFT)),
            Paragraph(str(h),                sty("rd",8,True,TA_CENTER,colors.HexColor("#1E6B2E"))),
            Paragraph(str(int(r.get("alpa",0))),  sty("rd",8,True,TA_CENTER,colors.HexColor("#C0281E"))),
            Paragraph(str(int(r.get("izin",0))),  sty("rd",8,True,TA_CENTER,colors.HexColor("#8A6000"))),
            Paragraph(str(int(r.get("sakit",0))), sty("rd",8,True,TA_CENTER,colors.HexColor("#1A4A9A"))),
            Paragraph(str(int(r.get("total_hari",0))), sty("rd",8,False,TA_CENTER)),
            Paragraph(pct,                   sty("rd",8,False,TA_CENTER)),
        ])

    rkp_tbl = Table(rkp_rows, colWidths=rkp_cols, repeatRows=1)
    rkp_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  hdr_c),
        ("FONTSIZE",      (0,0),(-1,-1), 8),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("BOX",           (0,0),(-1,-1), 0.5, thin),
        ("INNERGRID",     (0,0),(-1,-1), 0.3, thin),
        *[("BACKGROUND",(0,i),(-1,i), alt_c) for i in range(2,len(rkp_rows),2)],
    ]))
    story.append(rkp_tbl)

    story.append(Spacer(1,12))
    story.append(Paragraph(f"Dicetak: {date.today().strftime('%d %B %Y')}", sty("ft",8,False,TA_RIGHT,colors.HexColor("#9aa0b8"))))

    doc.build(story)
    return buf.getvalue()


def generate_absen_siswa_pdf(
    kelas: str,
    bulan: int,
    tahun: int,
    df_harian: pd.DataFrame,
    df_rekap:  pd.DataFrame,
) -> bytes:
    """PDF absensi siswa per kelas per bulan — A4 landscape."""
    import calendar as cal_mod
    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import PageBreak

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm)

    nama_bulan   = cal_mod.month_name[bulan]
    jumlah_hari  = cal_mod.monthrange(tahun, bulan)[1]
    hdr_c = colors.HexColor("#1F4E79")
    alt_c = colors.HexColor("#EBF5FB")
    thin  = colors.HexColor("#cccccc")

    def sty(name, size=8, bold=False, align=TA_CENTER, color=BLACK):
        return ParagraphStyle(name, fontName="Helvetica-Bold" if bold else "Helvetica",
                              fontSize=size, textColor=color, alignment=align,
                              leading=size*1.3, spaceAfter=0, spaceBefore=0)

    # Build status map
    status_map = {}
    if not df_harian.empty:
        for _, r in df_harian.iterrows():
            status_map[(r.get("nama",""), str(r.get("tanggal","")))] = r.get("status","")

    story = []

    # Halaman 1: Harian
    story.append(Paragraph(f"DAFTAR HADIR SISWA — {kelas.upper()}", sty("T",10,True)))
    story.append(Paragraph(f"{nama_bulan.upper()} {tahun} | {NAMA_SEKOLAH}", sty("S",9)))
    story.append(Spacer(1,6))

    col_w = [0.8*cm, 3.5*cm] + [0.52*cm]*jumlah_hari + [0.65*cm]*4
    hdr   = [Paragraph("No",sty("h",7,True,TA_CENTER,colors.white)),
             Paragraph("Nama Siswa",sty("h",7,True,TA_LEFT,colors.white))] + \
            [Paragraph(str(d),sty("h",6,True,TA_CENTER,colors.white)) for d in range(1,jumlah_hari+1)] + \
            [Paragraph(t,sty("h",7,True,TA_CENTER,colors.white)) for t in ["H","A","I","S"]]

    rows = [hdr]
    for idx,(_, rec) in enumerate(df_rekap.iterrows(),1):
        nama = rec.get("nama","")
        row  = [Paragraph(str(idx),sty("d",6,False,TA_CENTER)),
                Paragraph(nama[:22],sty("d",6,False,TA_LEFT))]
        for d in range(1,jumlah_hari+1):
            tgl = f"{tahun}-{str(bulan).zfill(2)}-{str(d).zfill(2)}"
            st  = status_map.get((nama,tgl),"")
            col = {"H":colors.HexColor("#1E6B2E"),"A":colors.HexColor("#C0281E"),
                   "I":colors.HexColor("#8A6000"),"S":colors.HexColor("#1A4A9A")}.get(st,BLACK)
            row.append(Paragraph(st,sty("d",6,False,TA_CENTER,col)))
        row += [Paragraph(str(int(rec.get(k,0))),sty("d",6,True,TA_CENTER)) for k in ["hadir","alpa","izin","sakit"]]
        rows.append(row)

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),hdr_c),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
        ("BOX",(0,0),(-1,-1),0.5,thin),("INNERGRID",(0,0),(-1,-1),0.3,thin),
        *[("BACKGROUND",(0,i),(-1,i),alt_c) for i in range(2,len(rows),2)],
    ]))
    story.append(tbl)

    # Halaman 2: Rekap
    story.append(PageBreak())
    story.append(Paragraph(f"REKAPITULASI KEHADIRAN SISWA — {kelas.upper()}", sty("T2",10,True)))
    story.append(Paragraph(f"{nama_bulan.upper()} {tahun} | {NAMA_SEKOLAH}", sty("S2",9)))
    story.append(Spacer(1,8))

    rkp_cols = [1*cm,5*cm,1.8*cm,1.8*cm,1.8*cm,1.8*cm,2*cm,2.5*cm]
    rkp_hdr  = [Paragraph(t,sty("rh",8,True,TA_CENTER,colors.white))
                for t in ["No","Nama Siswa","Hadir","Alpa","Izin","Sakit","Total Hari","% Hadir"]]
    rkp_rows = [rkp_hdr]
    for idx,(_, r) in enumerate(df_rekap.iterrows(),1):
        total = r.get("total_hari",0) or 1
        h     = int(r.get("hadir",0))
        pct   = f"{h/total*100:.1f}%"
        rkp_rows.append([
            Paragraph(str(idx),sty("rd",8,False,TA_CENTER)),
            Paragraph(r.get("nama","")[:25],sty("rd",8,False,TA_LEFT)),
            Paragraph(str(h),sty("rd",8,True,TA_CENTER,colors.HexColor("#1E6B2E"))),
            Paragraph(str(int(r.get("alpa",0))),sty("rd",8,True,TA_CENTER,colors.HexColor("#C0281E"))),
            Paragraph(str(int(r.get("izin",0))),sty("rd",8,True,TA_CENTER,colors.HexColor("#8A6000"))),
            Paragraph(str(int(r.get("sakit",0))),sty("rd",8,True,TA_CENTER,colors.HexColor("#1A4A9A"))),
            Paragraph(str(int(r.get("total_hari",0))),sty("rd",8,False,TA_CENTER)),
            Paragraph(pct,sty("rd",8,False,TA_CENTER)),
        ])

    rkp_tbl = Table(rkp_rows, colWidths=rkp_cols, repeatRows=1)
    rkp_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),hdr_c),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("BOX",(0,0),(-1,-1),0.5,thin),("INNERGRID",(0,0),(-1,-1),0.3,thin),
        *[("BACKGROUND",(0,i),(-1,i),alt_c) for i in range(2,len(rkp_rows),2)],
    ]))
    story.append(rkp_tbl)
    story.append(Spacer(1,12))
    story.append(Paragraph(f"Dicetak: {date.today().strftime('%d %B %Y')}",
                            sty("ft",8,False,TA_RIGHT,colors.HexColor("#9aa0b8"))))

    doc.build(story)
    return buf.getvalue()


def generate_jurnal_pdf(
    guru_nama: str,
    kelas: str,
    bulan: int,
    tahun: int,
    df_jurnal: pd.DataFrame,
) -> bytes:
    """PDF jurnal harian guru per bulan — A4 portrait."""
    import calendar as cal_mod
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm)

    nama_bulan = cal_mod.month_name[bulan]
    hdr_c = colors.HexColor("#1F4E79")
    alt_c = colors.HexColor("#EBF5FB")
    thin  = colors.HexColor("#cccccc")
    W     = A4[0] - 3*cm

    def sty(name, size=9, bold=False, align=TA_LEFT, color=BLACK):
        return ParagraphStyle(name, fontName="Helvetica-Bold" if bold else "Helvetica",
                              fontSize=size, textColor=color, alignment=align,
                              leading=size*1.4, spaceAfter=0, spaceBefore=0)

    story = []
    story.append(Paragraph(f"JURNAL HARIAN MENGAJAR", sty("T",12,True,TA_CENTER)))
    story.append(Paragraph(f"{guru_nama}", sty("S",10,False,TA_CENTER)))
    story.append(Paragraph(f"Kelas: {kelas} | {nama_bulan} {tahun} | {NAMA_SEKOLAH}", sty("S2",9,False,TA_CENTER)))
    story.append(Spacer(1,8))

    col_w = [1.2*cm, 2*cm, 2.5*cm, 2.5*cm, 5.5*cm, 3.3*cm]
    hdr   = [Paragraph(t,sty("h",8,True,TA_CENTER,colors.white))
             for t in ["No","Tanggal","Mapel","Topik","Aktivitas Pembelajaran","Catatan"]]
    rows  = [hdr]
    for idx,(_, j) in enumerate(df_jurnal.iterrows(),1):
        rows.append([
            Paragraph(str(idx),sty("d",8,False,TA_CENTER)),
            Paragraph(str(j.get("tanggal","")),sty("d",8)),
            Paragraph(str(j.get("mapel","")),sty("d",8)),
            Paragraph(str(j.get("topik","")),sty("d",8)),
            Paragraph(str(j.get("aktivitas","")),sty("d",8)),
            Paragraph(str(j.get("catatan","")),sty("d",8)),
        ])

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),hdr_c),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),12),
        ("BOX",(0,0),(-1,-1),0.5,thin),("INNERGRID",(0,0),(-1,-1),0.3,thin),
        *[("BACKGROUND",(0,i),(-1,i),alt_c) for i in range(2,len(rows),2)],
    ]))
    story.append(tbl)
    story.append(Spacer(1,12))
    story.append(Paragraph(f"Dicetak: {date.today().strftime('%d %B %Y')}",
                            sty("ft",8,False,TA_RIGHT,colors.HexColor("#9aa0b8"))))
    doc.build(story)
    return buf.getvalue()
